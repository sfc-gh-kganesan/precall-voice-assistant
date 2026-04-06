from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import yaml

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ExcelParseError(Exception):
    pass


@dataclass
class ColumnMetadata:
    name: str
    description: str = ""
    synonyms: list[str] = field(default_factory=list)
    glossary: str = ""
    comments: str = ""


@dataclass
class TableMetadata:
    name: str
    domain: str = ""
    purpose: str = ""
    usage: str = ""
    description: str = ""
    columns: dict[str, ColumnMetadata] = field(default_factory=dict)


@dataclass
class ExcelMetadata:
    source_file: str
    domains: dict[str, dict[str, TableMetadata]] = field(default_factory=dict)


def _ensure_openpyxl() -> None:
    if openpyxl is None:
        raise ExcelParseError(
            "openpyxl is required for Excel parsing. Install with: pip install openpyxl"
        )


def _normalize_domain(domain: str) -> str:
    return domain.lower().strip()


def _sanitize_text(value: str | None) -> str:
    if not value:
        return ""
    replacements = {
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2013": "-",
        "\u2014": "-",
        "\u2026": "...",
        "\u2022": "*",
        "\u00a0": " ",
        "\u2032": "'",
        "\u2033": '"',
        "\u00b4": "'",
        "\u0092": "'",
        "\u0091": "'",
        "\u0093": '"',
        "\u0094": '"',
        "\u0096": "-",
        "\u0097": "-",
        "\u0085": "...",
        "\u0095": "*",
    }
    result = value
    for old, new in replacements.items():
        result = result.replace(old, new)
    result = result.encode("ascii", errors="ignore").decode("ascii")
    return result


def _parse_synonyms(value: str | None) -> list[str]:
    if not value:
        return []
    sanitized = _sanitize_text(value)
    raw = sanitized.replace("\n", ",")
    return [s.strip() for s in raw.split(",") if s.strip()]


def _extract_domain_from_sheet_name(sheet_name: str, prefix: str) -> str | None:
    if not sheet_name.startswith(prefix):
        return None
    domain = sheet_name[len(prefix) :].strip()
    return domain if domain else None


def _parse_column_list_sheet(ws: Any, domain: str, metadata: ExcelMetadata) -> None:
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())
    col_map = {
        "table": None,
        "column": None,
        "description": None,
        "synonyms": None,
        "glossary": None,
        "comments": None,
    }
    header_aliases = {
        "table": ["db object name", "table name", "table", "object name"],
        "column": ["column name", "column", "field name", "field"],
        "description": ["column description", "description", "desc", "column desc"],
        "synonyms": ["synonyms", "synonym", "aliases", "alias"],
        "glossary": ["glossary", "glossary term", "term"],
        "comments": ["comments", "comment", "notes", "note"],
    }
    for field_name, aliases in header_aliases.items():
        for i, header in enumerate(headers):
            if header in aliases:
                col_map[field_name] = i
                break
    if col_map["table"] is None or col_map["column"] is None:
        return
    domain_key = _normalize_domain(domain)
    if domain_key not in metadata.domains:
        metadata.domains[domain_key] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map["table"]]:
            continue
        table_name = str(row[col_map["table"]]).strip().upper()
        column_name = str(row[col_map["column"]] or "").strip().upper()
        if not table_name or not column_name:
            continue
        if table_name not in metadata.domains[domain_key]:
            metadata.domains[domain_key][table_name] = TableMetadata(
                name=table_name, domain=domain_key
            )
        table_meta = metadata.domains[domain_key][table_name]
        description = ""
        if col_map["description"] is not None and row[col_map["description"]]:
            description = _sanitize_text(str(row[col_map["description"]]).strip())
        synonyms = []
        if col_map["synonyms"] is not None:
            synonyms = _parse_synonyms(row[col_map["synonyms"]])
        glossary = ""
        if col_map["glossary"] is not None and row[col_map["glossary"]]:
            glossary = _sanitize_text(str(row[col_map["glossary"]]).strip())
        comments = ""
        if col_map["comments"] is not None and row[col_map["comments"]]:
            comments = _sanitize_text(str(row[col_map["comments"]]).strip())
        table_meta.columns[column_name] = ColumnMetadata(
            name=column_name,
            description=description,
            synonyms=synonyms,
            glossary=glossary,
            comments=comments,
        )


def _parse_table_desc_sheet(ws: Any, domain: str, metadata: ExcelMetadata) -> None:
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())
    col_map = {"domain": None, "table": None, "purpose": None, "usage": None, "description": None}
    header_aliases = {
        "domain": ["domain", "business domain", "area"],
        "table": ["table name", "table", "object name", "db object name"],
        "purpose": ["purpose", "table purpose"],
        "usage": ["usage", "table usage", "use case"],
        "description": [
            "longform description",
            "description",
            "long description",
            "detailed description",
        ],
    }
    for field_name, aliases in header_aliases.items():
        for i, header in enumerate(headers):
            if header in aliases:
                col_map[field_name] = i
                break
    if col_map["table"] is None:
        return
    domain_key = _normalize_domain(domain)
    if domain_key not in metadata.domains:
        metadata.domains[domain_key] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map["table"]]:
            continue
        table_name = str(row[col_map["table"]]).strip().upper()
        if not table_name:
            continue
        if table_name not in metadata.domains[domain_key]:
            metadata.domains[domain_key][table_name] = TableMetadata(
                name=table_name, domain=domain_key
            )
        table_meta = metadata.domains[domain_key][table_name]
        if col_map["purpose"] is not None and row[col_map["purpose"]]:
            table_meta.purpose = _sanitize_text(str(row[col_map["purpose"]]).strip())
        if col_map["usage"] is not None and row[col_map["usage"]]:
            table_meta.usage = _sanitize_text(str(row[col_map["usage"]]).strip())
        if col_map["description"] is not None and row[col_map["description"]]:
            table_meta.description = _sanitize_text(str(row[col_map["description"]]).strip())


def parse_excel(excel_path: str | Path) -> ExcelMetadata:
    _ensure_openpyxl()
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise ExcelParseError(f"Excel file not found: {excel_path}")
    if excel_path.suffix.lower() not in (".xlsx", ".xls"):
        raise ExcelParseError(f"Expected Excel file (.xlsx/.xls), got: {excel_path.suffix}")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        raise ExcelParseError(f"Failed to open Excel file: {e}") from e
    metadata = ExcelMetadata(source_file=excel_path.name)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.startswith("Column List"):
            for prefix in ["Column List - ", "Column List "]:
                domain = _extract_domain_from_sheet_name(sheet_name, prefix)
                if domain:
                    _parse_column_list_sheet(ws, domain, metadata)
                    break
        elif sheet_name.startswith("Table Desc"):
            for prefix in ["Table Desc - ", "Table Desc "]:
                domain = _extract_domain_from_sheet_name(sheet_name, prefix)
                if domain:
                    _parse_table_desc_sheet(ws, domain, metadata)
                    break
    wb.close()
    if not metadata.domains:
        raise ExcelParseError(
            "No valid metadata sheets found. Expected sheets named 'Column List - {Domain}' or 'Table Desc {Domain}'"
        )
    return metadata


def get_column_metadata(
    metadata: ExcelMetadata, domain: str, table_name: str, column_name: str
) -> ColumnMetadata | None:
    domain_key = _normalize_domain(domain)
    table_key = table_name.upper()
    column_key = column_name.upper()
    domain_tables = metadata.domains.get(domain_key, {})
    table_meta = domain_tables.get(table_key)
    if table_meta is None:
        return None
    return table_meta.columns.get(column_key)


def get_table_metadata(
    metadata: ExcelMetadata, domain: str, table_name: str
) -> TableMetadata | None:
    domain_key = _normalize_domain(domain)
    table_key = table_name.upper()
    domain_tables = metadata.domains.get(domain_key, {})
    return domain_tables.get(table_key)


def to_yaml(metadata: ExcelMetadata) -> str:
    def obj_to_dict(obj: Any) -> Any:
        if hasattr(obj, "__dataclass_fields__"):
            return {k: obj_to_dict(v) for k, v in asdict(obj).items()}
        elif isinstance(obj, dict):
            return {k: obj_to_dict(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [obj_to_dict(item) for item in obj]
        else:
            return obj

    data = obj_to_dict(metadata)
    return yaml.dump(data, default_flow_style=False, sort_keys=False, allow_unicode=True)
